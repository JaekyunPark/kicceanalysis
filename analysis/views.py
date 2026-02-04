from django.http import JsonResponse
from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from surveys.models import SurveyData, Codebook
import pandas as pd
import numpy as np
from scipy import stats
import json
from django.core.serializers.json import DjangoJSONEncoder
from .models import AnalysisPreset
try:
    import pyreadstat
except ImportError:
    pyreadstat = None

def dataframe_to_template_dict(df):
    """Pandas DataFrame을 템플릿에서 사용하기 쉬운 딕셔너리로 변환"""
    return {
        'columns': list(df.columns),
        'index': list(df.index),
        'data': df.values.tolist()
    }

def weighted_mean(values, weights):
    """가중 평균 계산"""
    try:
        return np.average(values, weights=weights)
    except ZeroDivisionError:
        return 0

def weighted_std(values, weights):
    """가중 표준편차 계산"""
    try:
        average = np.average(values, weights=weights)
        variance = np.average((values - average)**2, weights=weights)
        # 표본 보정 (Frequency weights assumption: sum(weights) is N)
        # Unbiased estimator: V * N / (N - 1)
        n = weights.sum()
        if n > 1:
            variance = variance * n / (n - 1)
        return np.sqrt(variance)
    except:
        return 0

def weighted_quantile(values, weights, quantile):
    """가중 분위수 계산"""
    try:
        df = pd.DataFrame({'val': values, 'weight': weights})
        df = df.sort_values('val')
        cumsum = df['weight'].cumsum()
        cutoff = df['weight'].sum() * quantile
        return df[cumsum >= cutoff]['val'].iloc[0]
    except:
        return 0

def weighted_ttest(group1, group2, w1, w2):
    """가중 t-test (Welch's t-test with effective degrees of freedom)"""
    try:
        n1 = w1.sum()
        n2 = w2.sum()
        
        m1 = weighted_mean(group1, w1)
        m2 = weighted_mean(group2, w2)
        
        v1 = weighted_std(group1, w1)**2
        v2 = weighted_std(group2, w2)**2
        
        # Welch's t-statistic
        se = np.sqrt(v1/n1 + v2/n2)
        if se == 0: return None
        
        t_stat = (m1 - m2) / se
        
        # Welch-Satterthwaite degrees of freedom
        df = (v1/n1 + v2/n2)**2 / ((v1/n1)**2/(n1-1) + (v2/n2)**2/(n2-1))
        
        # Two-sided p-value
        p_val = stats.t.sf(np.abs(t_stat), df) * 2
        return t_stat, p_val, df
    except:
        return None

def weighted_anova(groups, weights_list):
    """가중 ANOVA (F-test)"""
    try:
        # 1. Grand Mean
        all_values = np.concatenate(groups)
        all_weights = np.concatenate(weights_list)
        grand_mean = weighted_mean(all_values, all_weights)
        
        # 2. Between Group Sum of Squares (SSB)
        ssb = 0
        for g, w in zip(groups, weights_list):
            m = weighted_mean(g, w)
            n = w.sum()
            ssb += n * (m - grand_mean)**2
            
        # 3. Within Group Sum of Squares (SSW)
        ssw = 0
        for g, w in zip(groups, weights_list):
            m = weighted_mean(g, w)
            ssw += np.sum(w * (g - m)**2)
            
        # 4. Degrees of Freedom
        k = len(groups)
        N = all_weights.sum()
        df_between = k - 1
        df_within = N - k
        
        if df_between <= 0 or df_within <= 0:
            return None
            
        # 5. Mean Squares
        msb = ssb / df_between
        msw = ssw / df_within
        
        if msw == 0: return None
        
        # 6. F-statistic
        f_stat = msb / msw
        p_val = stats.f.sf(f_stat, df_between, df_within)
        
        return f_stat, p_val, df_between, df_within
    except:
        return None

    """DataFrame을 템플릿에서 사용 가능한 딕셔너리로 변환"""
    return {
        'index': list(df.index),
        'columns': list(df.columns),
        'data': df.values.tolist()
    }

def format_test_statistic(test_value, df, p_value, test_format, test_type='chi'):
    """검정 통계량을 사용자 지정 형식으로 포맷팅
    
    Args:
        test_value: 검정 통계량 값 (χ², t, F 등)
        df: 자유도 (단일 값 또는 튜플)
        p_value: 유의확률
        test_format: 표시 옵션 딕셔너리
            - show_label: 레이블 표시 여부 (χ² = , t = , F = )
            - show_value: 검정값 표시 여부
            - show_df: 자유도 표시 여부
            - show_stars: 유의수준 별표 표시 여부
            - show_pvalue: p-value 표시 여부
        test_type: 검정 유형 ('chi', 't', 'f')
    
    Returns:
        포맷팅된 문자열
    """
    result = ''
    
    # 1. 레이블
    if test_format.get('show_label', True):
        if test_type == 'chi':
            result += 'χ² = '
        elif test_type == 't':
            result += 't = '
        elif test_type == 'f':
            result += 'F = '
    
    # 2. 검정값
    if test_format.get('show_value', True):
        result += f'{test_value:.3f}'
    
    # 3. 자유도
    if test_format.get('show_df', True):
        if isinstance(df, tuple):
            # ANOVA의 경우 (df_between, df_within)
            result += f'({df[0]}, {df[1]})'
        else:
            # chi-square, t-test의 경우
            result += f'({df})'
    
    # 4. 유의수준 별표
    if test_format.get('show_stars', True):
        if p_value < 0.001:
            result += '***'
        elif p_value < 0.01:
            result += '**'
        elif p_value < 0.05:
            result += '*'
    
    # 5. p-value
    if test_format.get('show_pvalue', False):
        result += f'({p_value:.3f})'
    
    return result if result else '-'

@login_required
def unified_analysis(request):
    """통합 분석 메인 페이지"""
    datasets = SurveyData.objects.filter(user=request.user)
    
    # GET 파라미터로 dataset_id가 넘어오면 미리 선택
    selected_dataset_id = request.GET.get('dataset_id')
    try:
        if selected_dataset_id:
            selected_dataset_id = int(selected_dataset_id)
    except (ValueError, TypeError):
        selected_dataset_id = None

    if request.method == 'POST':
        dataset_id = request.POST.get('dataset')
        codebook_id = request.POST.get('codebook')
        row_variables = request.POST.get('row_variables', '').split(',')
        col_variables = request.POST.get('col_variables', '').split(',')
        col_statistics = json.loads(request.POST.get('col_statistics', '{}'))
        weight_variable = request.POST.get('weight_variable', '')
        
        # 빈 문자열 제거
        row_variables = [v for v in row_variables if v]
        col_variables = [v for v in col_variables if v]
        
        if not dataset_id or not row_variables or not col_variables:
            return render(request, 'analysis/unified_analysis.html', {
                'datasets': datasets,
                'error': '필수 항목을 모두 입력해주세요.',
                'selected_dataset_id': int(dataset_id) if dataset_id else None
            })
        
        # 분석 수행
        result = perform_unified_analysis(
            dataset_id=dataset_id,
            codebook_id=codebook_id,
            row_variables=row_variables,
            col_variables=col_variables,
            col_statistics=col_statistics,
            weight_variable=weight_variable,
            user=request.user,
            options={
                'show_total': request.POST.get('show_total') == 'on',
                'auto_test': request.POST.get('auto_test') == 'on',
                'percent_symbol': request.POST.get('percent_symbol', 'yes') == 'yes',
                'mean_decimals': int(request.POST.get('mean_decimals', '2')),
                'chi_format': {
                    'show_label': request.POST.get('show_chi_label') == 'on',
                    'show_value': request.POST.get('show_chi_value') == 'on',
                    'show_df': request.POST.get('show_chi_df') == 'on',
                    'show_stars': request.POST.get('show_chi_stars') == 'on',
                    'show_pvalue': request.POST.get('show_chi_pvalue') == 'on',
                },
                'tf_format': {
                    'show_label': request.POST.get('show_tf_label') == 'on',
                    'show_value': request.POST.get('show_tf_value') == 'on',
                    'show_df': request.POST.get('show_tf_df') == 'on',
                    'show_stars': request.POST.get('show_tf_stars') == 'on',
                    'show_pvalue': request.POST.get('show_tf_pvalue') == 'on',
                }
            }
        )
        
        return render(request, 'analysis/analysis_result.html', {
            'result': result,
            'dataset_id': dataset_id,
            'codebook_id': codebook_id,
            'row_variables': row_variables,
            'col_variables': col_variables,
            'col_statistics': json.dumps(col_statistics),
            'weight_variable': weight_variable,
            'show_total': request.POST.get('show_total') == 'on',
            'auto_test': request.POST.get('auto_test') == 'on',
            'percent_symbol': request.POST.get('percent_symbol', 'yes'),
            'mean_decimals': request.POST.get('mean_decimals', '2'),
            'show_chi_label': request.POST.get('show_chi_label') == 'on',
            'show_chi_value': request.POST.get('show_chi_value') == 'on',
            'show_chi_df': request.POST.get('show_chi_df') == 'on',
            'show_chi_stars': request.POST.get('show_chi_stars') == 'on',
            'show_chi_pvalue': request.POST.get('show_chi_pvalue') == 'on',
            'show_tf_label': request.POST.get('show_tf_label') == 'on',
            'show_tf_value': request.POST.get('show_tf_value') == 'on',
            'show_tf_df': request.POST.get('show_tf_df') == 'on',
            'show_tf_stars': request.POST.get('show_tf_stars') == 'on',
            'show_tf_pvalue': request.POST.get('show_tf_pvalue') == 'on',
        })
    
    return render(request, 'analysis/unified_analysis.html', {
        'datasets': datasets,
        'selected_dataset_id': selected_dataset_id
    })

def perform_unified_analysis(dataset_id, codebook_id, row_variables, col_variables, 
                             col_statistics, user, options, weight_variable=None):
    """통합 분석 수행"""
    dataset = get_object_or_404(SurveyData, id=dataset_id, user=user)
    
    # 데이터 로드
    if dataset.file.name.lower().endswith('.csv'):
        df = pd.read_csv(dataset.file.path)
    elif dataset.file.name.lower().endswith('.sav'):
        df = pd.read_spss(dataset.file.path, convert_categoricals=False)
    else:
        df = pd.read_excel(dataset.file.path)
    
    # 코드북 로드 (선택사항)
    codebook_data = None
    if codebook_id:
        codebook = get_object_or_404(Codebook, id=codebook_id, user=user)
        codebook_data = load_codebook(codebook)
    elif dataset.codebooks.exists():
        # 코드북이 선택되지 않았지만 연결된 코드북이 있는 경우 자동 사용 (특히 SAV 파일)
        codebook = dataset.codebooks.first()
        codebook_data = load_codebook(codebook)
    
    results = []
    
    # 열 변수별로 그룹화하여 분석 수행
    for col_var in col_variables:
        stats_config = col_statistics.get(col_var, {})
        col_label = get_variable_label(col_var, codebook_data) if codebook_data else col_var
        
        # 통계량 유형 판단
        has_frequency = stats_config.get('count') or stats_config.get('row_pct') or stats_config.get('col_pct')
        has_descriptive = any([stats_config.get('mean'), stats_config.get('std'), 
                              stats_config.get('median'), stats_config.get('min'), 
                              stats_config.get('max'), stats_config.get('q1'), 
                              stats_config.get('q3')])
        
        analysis_result = {
            'row_variables': row_variables,
            'col_var': col_var,
            'row_labels': [get_variable_label(rv, codebook_data) if codebook_data else rv for rv in row_variables],
            'col_label': col_label,
            'stats_config': stats_config,
        }
        
        # 빈도와 기술통계가 모두 선택된 경우 - 통합 테이블
        if has_frequency and has_descriptive:
            unified_result = perform_combined_freq_desc_analysis(
                df, row_variables, col_var, stats_config,
                codebook_data, options, weight_variable
            )
            analysis_result['unified'] = unified_result
        else:
            # 빈도 분석만
            if has_frequency:
                freq_result = perform_frequency_analysis_combined(
                    df, row_variables, col_var, stats_config, 
                    codebook_data, options, weight_variable
                )
                analysis_result['frequency'] = freq_result
            
            # 기술통계 분석만
            if has_descriptive:
                desc_result = perform_descriptive_analysis_combined(
                    df, row_variables, col_var, stats_config, 
                    codebook_data, options, weight_variable
                )
                analysis_result['descriptive'] = desc_result
        
        results.append(analysis_result)
    
    return results

def weighted_ttest_pooled(group1, group2, w1, w2):
    """가중 등분산 가정 t-test (Pooled t-test)"""
    try:
        n1 = w1.sum()
        n2 = w2.sum()
        
        m1 = weighted_mean(group1, w1)
        m2 = weighted_mean(group2, w2)
        
        v1 = weighted_std(group1, w1)**2
        v2 = weighted_std(group2, w2)**2
        
        # Pooled Variance
        # S_p^2 = ((n1-1)v1 + (n2-1)v2) / (n1+n2-2)
        df = n1 + n2 - 2
        sp_sq = ((n1 - 1)*v1 + (n2 - 1)*v2) / df
        
        se = np.sqrt(sp_sq * (1/n1 + 1/n2))
        if se == 0: return None
        
        t_stat = (m1 - m2) / se
        
        # Two-sided p-value
        p_val = stats.t.sf(np.abs(t_stat), df) * 2
        return t_stat, p_val, df
    except:
        return None

def weighted_levene(groups, weights_list):
    """가중 Levene 검정 (Weighted Levene's Test) - using Mean"""
    try:
        # 1. 각 그룹의 가중 평균 계산
        group_means = []
        for g, w in zip(groups, weights_list):
            group_means.append(weighted_mean(g, w))
            
        # 2. 절대 편차(Absolute Deviations) 계산
        deviations = []
        dev_weights = []
        
        for i, (g, w) in enumerate(zip(groups, weights_list)):
            mean = group_means[i]
            dev = np.abs(g - mean)
            deviations.append(dev)
            dev_weights.append(w)
            
        # 3. 편차들에 대해 Weighted ANOVA 수행
        return weighted_anova(deviations, dev_weights)
    except:
        return None

def perform_combined_freq_desc_analysis(df, row_variables, col_var, stats_config, codebook_data, options, weight_variable=None):
    """빈도 분석과 기술통계를 하나의 테이블로 통합"""
    
    # 가중치 변수 확인
    weights = None
    if weight_variable and weight_variable in df.columns:
        weights = pd.to_numeric(df[weight_variable], errors='coerce').fillna(0)
    
    # 열 변수를 연속형으로 변환 시도 (기술통계용)
    try:
        df_copy = df.copy()
        df_copy[col_var] = pd.to_numeric(df_copy[col_var], errors='coerce')
    except:
        pass
    
    col_label = get_variable_label(col_var, codebook_data) if codebook_data else col_var
    
    # 빈도 분석 열 변수의 값들 (열 헤더)
    freq_columns = []
    if stats_config.get('count') or stats_config.get('row_pct') or stats_config.get('col_pct'):
        # 열 변수의 고유값들
        unique_values = sorted(df[col_var].dropna().unique())
        for val in unique_values:
            if codebook_data and col_var in codebook_data and 'values' in codebook_data[col_var]:
                value_labels = codebook_data[col_var]['values']
                try:
                    key = str(int(float(val)))
                except:
                    key = str(val)
                label = value_labels.get(key, str(val))
            else:
                label = str(val)
            freq_columns.append(label)
        
        if options['show_total']:
            freq_columns.append('Total')
        
        # 행% 일 때만 계와 사례수 추가
        if stats_config.get('row_pct'):
            freq_columns.append('계')
            freq_columns.append('사례수')
    
    # 기술통계 열 (통계량 이름)
    desc_columns = []
    stats_to_calc = []
    
    if stats_config.get('mean'):
        if weights is not None:
             stats_to_calc.append(('평균', 'mean', lambda x, w: weighted_mean(x, w)))
        else:
            stats_to_calc.append(('평균', 'mean', lambda x: x.mean()))
        desc_columns.append(f"{col_label} (평균)")
    if stats_config.get('median'):
        if weights is not None:
            stats_to_calc.append(('중앙값', 'median', lambda x, w: weighted_median(x, w) if hasattr(weighted_mean, '__code__') else weighted_quantile(x, w, 0.5))) # weighted_median wrapper
        else:
            stats_to_calc.append(('중앙값', 'median', lambda x: x.median()))
        desc_columns.append(f"{col_label} (중앙값)")
    if stats_config.get('std'):
        if weights is not None:
            stats_to_calc.append(('표준편차', 'std', lambda x, w: weighted_std(x, w)))
        else:
            stats_to_calc.append(('표준편차', 'std', lambda x: x.std()))
        desc_columns.append(f"{col_label} (표준편차)")
    if stats_config.get('min'):
        stats_to_calc.append(('최솟값', 'min', lambda x: x.min())) # Weighted Min is same as Unweighted
        desc_columns.append(f"{col_label} (최솟값)")
    if stats_config.get('max'):
        stats_to_calc.append(('최댓값', 'max', lambda x: x.max())) # Weighted Max is same as Unweighted
        desc_columns.append(f"{col_label} (최댓값)")
    if stats_config.get('q1'):
        if weights is not None:
            stats_to_calc.append(('Q1 (25%)', 'q1', lambda x, w: weighted_quantile(x, w, 0.25)))
        else:
            stats_to_calc.append(('Q1 (25%)', 'q1', lambda x: x.quantile(0.25)))
        desc_columns.append(f"{col_label} (Q1)")
    if stats_config.get('q3'):
        if weights is not None:
            stats_to_calc.append(('Q3 (75%)', 'q3', lambda x, w: weighted_quantile(x, w, 0.75)))
        else:
            stats_to_calc.append(('Q3 (75%)', 'q3', lambda x: x.quantile(0.75)))
        desc_columns.append(f"{col_label} (Q3)")
    
    # 전체 열 헤더
    all_columns = freq_columns + desc_columns
    
    # 각 행 변수별로 처리
    row_var_results = []
    
    for row_var in row_variables:
        row_label = get_variable_label(row_var, codebook_data) if codebook_data else row_var
        rows = []
        
        # 이 행 변수의 각 그룹에 대해
        unique_row_vals = sorted(df[row_var].dropna().unique())
        
        for row_val in unique_row_vals:
            # 값 레이블 적용
            if codebook_data and row_var in codebook_data and 'values' in codebook_data[row_var]:
                value_labels = codebook_data[row_var]['values']
                try:
                    key = str(int(float(row_val)))
                except:
                    key = str(row_val)
                row_val_label = value_labels.get(key, str(row_val))
            else:
                row_val_label = str(row_val)
            
            row_name = f"{row_label}: {row_val_label}"
            row_data = []
            
            # 빈도 분석 데이터
            if freq_columns:
                subset = df[df[row_var] == row_val]
                crosstab = pd.crosstab(subset[row_var], subset[col_var])
                
                # 행 백분율 계산 (선택된 경우)
                if stats_config.get('row_pct'):
                    if weights is not None:
                         # 가중치 적용 교차표
                         crosstab_weighted = pd.crosstab(subset[row_var], subset[col_var], values=weights[subset.index], aggfunc='sum').fillna(0)
                         row_pct = crosstab_weighted.div(crosstab_weighted.sum(axis=1), axis=0) * 100
                         row_count = int(round(crosstab_weighted.sum().sum()))
                    else:
                        row_pct = crosstab.div(crosstab.sum(axis=1), axis=0) * 100
                        row_count = int(crosstab.sum().sum())  # 사례수
                    
                    # 백분율 포맷 결정
                    show_percent = options.get('percent_symbol', True)
                    
                    # 원본 백분율 값을 저장 (계 계산용)
                    pct_values = []
                    
                    # 각 열 값에 대한 백분율 및 빈도 병합 (사용자 요청: 빈도(비율) 형식)
                    for col_val in unique_values:
                        if col_val in row_pct.columns:
                            pct_val = row_pct.iloc[0, row_pct.columns.get_loc(col_val)]
                            pct_values.append(pct_val)
                            
                            if weights is not None:
                                count_val = crosstab_weighted.iloc[0, crosstab_weighted.columns.get_loc(col_val)]
                            else:
                                count_val = crosstab.iloc[0, crosstab.columns.get_loc(col_val)]
                            
                            # 빈도(비율) 형식으로 저장
                            row_data.append(f"{int(round(count_val))}({pct_val:.1f})")
                        else:
                            pct_values.append(0.0)
                            row_data.append("0(0.0)")
                    
                    if options['show_total']:
                        row_data.append(f"{row_count}(100.0)")
                    
                    # 계 추가
                    row_sum = sum(pct_values)
                    row_data.append(f"{row_count}({row_sum:.1f})")
                    
                    # 사례수 추가
                    row_data.append(f"({row_count})")
                        
                elif stats_config.get('count'):
                    # 빈도
                    for col_val in unique_values:
                        if col_val in crosstab.columns:
                            if weights is not None:
                                # 가중치 적용된 빈도
                                w_subset = subset[subset[col_var] == col_val]
                                w_sum = weights[w_subset.index].sum()
                                row_data.append(round(w_sum, 1)) # 가중치는 소수점 가능
                            else:
                                row_data.append(int(crosstab.iloc[0, crosstab.columns.get_loc(col_val)]))
                        else:
                            row_data.append(0)
                    
                    if options['show_total']:
                        if weights is not None:
                            row_data.append(round(weights[subset.index].sum(), 1))
                        else:
                            row_data.append(int(crosstab.sum().sum()))
            
            # 기술통계 데이터
            if desc_columns and 'df_copy' in locals():
                subset_numeric = df_copy[df_copy[row_var] == row_val][col_var].dropna()
                
                mean_decimals = options.get('mean_decimals', 2)
                
                for stat_label, stat_type, func in stats_to_calc:
                    if weights is not None:
                        if stat_type in ['mean', 'median', 'std', 'q1', 'q3']:
                            target_weights = weights[subset_numeric.index]
                            value = func(subset_numeric, target_weights)
                        else:
                            value = func(subset_numeric)
                    else:
                        value = func(subset_numeric)
                    
                    # 평균, 중앙값, 표준편차 등 소수점 적용
                    if stat_type in ['mean', 'median', 'std', 'q1', 'q3']:
                        row_data.append(round(value, mean_decimals))
                    else:
                        # 최솟값, 최댓값은 정수일 수 있음
                        if isinstance(value, (int, float, np.integer, np.floating)):
                             if stat_type in ['min', 'max'] and isinstance(value, float) and value.is_integer():
                                 row_data.append(int(value))
                             else:
                                 row_data.append(round(value, mean_decimals) if isinstance(value, float) else value)
                        else:
                            row_data.append(value)
            
            rows.append({
                'label': row_name,
                'values': row_data
            })
        
        # Total 행 추가
        if options['show_total']:
            total_data = []
            
            # 빈도 분석 Total
            if freq_columns:
                if stats_config.get('row_pct'):
                    # 전체 백분율
                    total_crosstab = pd.crosstab(df[row_var], df[col_var], normalize='all') * 100
                    col_totals = total_crosstab.sum(axis=0)
                    total_count = len(df[row_var].dropna())  # 전체 사례수
                    
                    for col_val in unique_values:
                        if col_val in col_totals.index:
                            # 전체 빈도 합계도 필요 (위에서 total_crosstab을 normalize='all'로 했으니 빈도 crosstab 다시 필요)
                            crosstab_count = pd.crosstab(df[row_var], df[col_var])
                            col_total_count = crosstab_count[col_val].sum()
                            total_data.append(f"{int(col_total_count)}({col_totals[col_val]:.1f})")
                        else:
                            total_data.append("0(0.0)")
                    
                    if options['show_total']:
                        total_data.append(f"{total_count}(100.0)")
                    
                    # 계 추가
                    pct_sum = sum([float(x.split('(')[1].rstrip(')')) for x in total_data if '(' in x])
                    total_data.append(f"{total_count}({pct_sum:.1f})")
                    
                    # 사례수 추가
                    total_data.append(f"({total_count})")
                        
                elif stats_config.get('count'):
                    total_crosstab = pd.crosstab(df[row_var], df[col_var])
                    col_totals = total_crosstab.sum(axis=0)
                    
                    for col_val in unique_values:
                        if col_val in col_totals.index:
                            total_data.append(int(col_totals[col_val]))
                        else:
                            total_data.append(0)
                    
                    if options['show_total']:
                        total_data.append(int(total_crosstab.sum().sum()))
            
            # 기술통계 Total
            if desc_columns and 'df_copy' in locals():
                mean_decimals = options.get('mean_decimals', 2)
                
                for stat_label, stat_type, func in stats_to_calc:
                    target_vals = df_copy[col_var].dropna()
                    if weights is not None:
                        if stat_type in ['mean', 'median', 'std', 'q1', 'q3']:
                            target_weights = weights[target_vals.index]
                            value = func(target_vals, target_weights)
                        else:
                            value = func(target_vals)
                    else:
                        value = func(target_vals)
                    
                    # 평균, 중앙값, 표준편차 등 소수점 적용
                    if stat_type in ['mean', 'median', 'std', 'q1', 'q3']:
                        total_data.append(round(value, mean_decimals))
                    else:
                         if isinstance(value, (int, float, np.integer, np.floating)):
                             if stat_type in ['min', 'max'] and isinstance(value, float) and value.is_integer():
                                 total_data.append(int(value))
                             else:
                                 total_data.append(round(value, mean_decimals) if isinstance(value, float) else value)
                         else:
                             total_data.append(value)
            
            rows.append({
                'label': 'Total',
                'values': total_data
            })
        
        # 통계 검정
        chi_square_test = None
        descriptive_test = None
        
        # 카이제곱 검정
        if options.get('auto_test') and (stats_config.get('row_pct') or stats_config.get('col_pct')):
            from scipy.stats import chi2_contingency
            crosstab_full = pd.crosstab(df[row_var], df[col_var])
            
            try:
                # 카이제곱 검정 (가중치 지원)
                if weights is not None:
                     # 가중치가 적용된 crosstab 사용
                     # SPSS 위변조 방지: 가중치 적용된 셀 빈도를 반올림하여 정수로 계산
                     crosstab_weighted = pd.crosstab(df[row_var], df[col_var], values=weights, aggfunc='sum').fillna(0).round()
                     chi2, p_value, dof, expected = chi2_contingency(crosstab_weighted, correction=False)
                else:
                    chi2, p_value, dof, expected = chi2_contingency(crosstab_full, correction=False)
                
                # 가중치가 있을 경우 검정 결과에 경고 표시 필요 (현재는 일단 단순 수행)
                if weights is not None:
                    pass 
                
                # 카이제곱 검정 포맷팅
                chi_format = options.get('chi_format', {
                    'show_value': True,
                    'show_df': True,
                    'show_stars': True,
                    'show_pvalue': False
                })
                formatted_stat = format_test_statistic(chi2, int(dof), p_value, chi_format, 'chi')
                
                chi_square_test = {
                    'chi2': chi2,
                    'p_value': p_value,
                    'dof': dof,
                    'test_name': 'Chi-square',
                    'formatted': formatted_stat
                }
            except:
                pass
        
        # t-test / ANOVA (평균 선택 시 수행)
        descriptive_test = None
        if options.get('auto_test') and stats_config.get('mean'):
            print(f"=== t/F 검정 시도 ===")
            print(f"auto_test: {options.get('auto_test')}")
            print(f"mean in stats_config: {stats_config.get('mean')}")
            
            try:
                # df_copy가 없으면 새로 생성
                if 'df_copy' not in locals():
                    print("df_copy 생성 중...")
                    df_copy = df.copy()
                    df_copy[col_var] = pd.to_numeric(df_copy[col_var], errors='coerce')
                    print(f"df_copy 생성 완료. 열 타입: {df_copy[col_var].dtype}")
                else:
                    print("df_copy 이미 존재")
                
                groups = [group[col_var].dropna() for name, group in df_copy.groupby(row_var)]
                print(f"그룹 개수: {len(groups)}")
                for i, g in enumerate(groups):
                    print(f"  그룹 {i}: {len(g)}개, 샘플: {g.head(3).tolist() if len(g) > 0 else 'empty'}")
                
                # t/F 검정 포맷팅
                tf_format = options.get('tf_format', {
                    'show_label': True,
                    'show_value': True,
                    'show_df': True,
                    'show_stars': True,
                    'show_pvalue': False
                })
                
                if len(groups) == 2:
                    try:
                        from scipy import stats as sp_stats
                        
                        # Levene 검정 결과에 따라 적절한 t-test 선택
                        if weights is not None:
                             # 가중치 t-test 수행
                             
                             # 리스트 준비
                             w1 = weights[groups[0].index]
                             w2 = weights[groups[1].index]
                             
                             # 1. 가중 Levene 검정
                             levene_result = weighted_levene(groups, [w1, w2])
                             levene_formatted_str = ""

                             if levene_result:
                                 f_levene, p_levene, df1_levene, df2_levene = levene_result
                                 levene_stat, levene_p = f_levene, p_levene
                                 # Levene 포맷팅
                                 levene_formatted_str = format_test_statistic(f_levene, (df1_levene, df2_levene), p_levene, tf_format, 'f')
                             else:
                                 levene_stat, levene_p = 0, 1 # 실패 시 등분산 가정
                                 
                             # 2. 가중 t-test (Welch)
                             t_welch, p_welch, df_welch_val = weighted_ttest(groups[0], groups[1], w1, w2)
                             
                             # 3. 가중 t-test (Pooled/Equal Variance)
                             t_pooled, p_pooled, df_pooled_val = weighted_ttest_pooled(groups[0], groups[1], w1, w2)
                             
                             # 결과 매핑
                             t_stat_equal, p_value_equal, df_equal = t_pooled, p_pooled, df_pooled_val
                             t_stat_welch, p_value_welch, df_welch = t_welch, p_welch, df_welch_val
                             
                             # 선택 로직
                             if levene_p < 0.05:
                                 t_stat, p_value, df_val = t_stat_welch, p_value_welch, df_welch
                                 test_name = "t-test (Welch)"
                             else:
                                 t_stat, p_value, df_val = t_stat_equal, p_value_equal, df_equal
                                 test_name = "t-test"
                        else:
                            # 기존 unweighted t-test 로직
                            
                            # Levene의 등분산 검정
                            levene_stat, levene_p = sp_stats.levene(groups[0], groups[1])
                            levene_formatted_str = f"F={levene_stat:.3f} ({levene_p:.3f})"
                            
                            # 등분산 가정 t-test (Student's t-test)
                            t_stat_equal, p_value_equal = sp_stats.ttest_ind(groups[0], groups[1], equal_var=True)
                            df_equal = len(groups[0]) + len(groups[1]) - 2
                            
                            # 등분산 가정하지 않음 t-test (Welch's t-test)
                            t_stat_welch, p_value_welch = sp_stats.ttest_ind(groups[0], groups[1], equal_var=False)
                            # Welch's t-test의 자유도 계산
                            n1, n2 = len(groups[0]), len(groups[1])
                            v1, v2 = groups[0].var(ddof=1), groups[1].var(ddof=1)
                            df_welch = ((v1/n1 + v2/n2)**2) / ((v1/n1)**2/(n1-1) + (v2/n2)**2/(n2-1))
                            
                            # p < 0.05이면 등분산 가정 위배 → Welch's t-test 사용
                            if levene_p < 0.05:
                                t_stat, p_value, df_val = t_stat_welch, p_value_welch, df_welch
                                test_name = "t-test (Welch)"
                            else:
                                t_stat, p_value, df_val = t_stat_equal, p_value_equal, df_equal
                                test_name = "t-test"
                        
                        formatted_stat = format_test_statistic(t_stat, df_val, p_value, tf_format, 't')
                        
                        print(f"Levene 검정: F={levene_stat:.3f}, p={levene_p:.3f}")
                        print(f"등분산 가정: t={t_stat_equal:.3f}, p={p_value_equal:.3f}, df={df_equal}")
                        print(f"Welch's t: t={t_stat_welch:.3f}, p={p_value_welch:.3f}, df={df_welch:.2f}")
                        print(f"선택된 검정: {test_name}, formatted: {formatted_stat}")
                        
                        descriptive_test = {
                            't_statistic': t_stat,
                            'p_value': p_value,
                            'test_name': test_name,
                            'df': df_val,
                            'levene_statistic': levene_stat,
                            'levene_p': levene_p,
                            't_equal': t_stat_equal,
                            'p_equal': p_value_equal,
                            'df_equal': df_equal,
                            't_welch': t_stat_welch,
                            'p_welch': p_value_welch,
                            'df_welch': df_welch,
                            'formatted': formatted_stat,
                            'levene_formatted': levene_formatted_str
                        }
                    except Exception as e:
                        print(f"t-test 실패: {e}")
                        import traceback
                        traceback.print_exc()
                        pass
                elif len(groups) >= 3:
                    try:
                        if weights is not None:
                             # 가중치 ANOVA
                             weights_list = [weights[g.index] for g in groups]
                             anova_result = weighted_anova(groups, weights_list)
                             
                             if anova_result:
                                 f_stat, p_value, df_between, df_within = anova_result
                             else:
                                 raise ValueError("Weighted ANOVA failed")
                        else:
                            from scipy import stats as sp_stats
                            f_stat, p_value = sp_stats.f_oneway(*groups)
                            df_between = len(groups) - 1
                            df_within = sum(len(g) for g in groups) - len(groups)
                        
                        formatted_stat = format_test_statistic(f_stat, (df_between, df_within), p_value, tf_format, 'f')
                        
                        print(f"ANOVA 성공: F={f_stat:.3f}, p={p_value:.3f}, df=({df_between}, {df_within})")
                        print(f"formatted: {formatted_stat}")
                        
                        descriptive_test = {
                            'f_statistic': f_stat,
                            'p_value': p_value,
                            'test_name': 'ANOVA (F-test)',
                            'df_between': df_between,
                            'df_within': df_within,
                            'formatted': formatted_stat
                        }
                    except Exception as e:
                        print(f"ANOVA 실패: {e}")
                        pass
                else:
                    print(f"검정 불가: 그룹 개수가 {len(groups)}개")
            except Exception as e:
                print(f"검정 통계 오류: {e}")
                import traceback
                traceback.print_exc()
                pass
        else:
            print(f"t/F 검정 건너뜀: auto_test={options.get('auto_test')}, mean={stats_config.get('mean')}")
        
        row_var_results.append({
            'rows': rows,
            'chi_square_test': chi_square_test,
            'descriptive_test': descriptive_test
        })
    
    return {
        'columns': all_columns,
        'freq_col_count': len(freq_columns),
        'desc_col_count': len(desc_columns),
        'row_var_results': row_var_results,
        'mean_decimals': options.get('mean_decimals', 2)
    }

def perform_frequency_analysis_combined(df, row_variables, col_var, stats_config, codebook_data, options, weight_variable=None):
    """여러 행 변수를 하나의 테이블로 통합한 빈도 분석"""
    
    # 가중치 변수 확인
    weights = None
    if weight_variable and weight_variable in df.columns:
        weights = pd.to_numeric(df[weight_variable], errors='coerce').fillna(0)
    
    # 각 행 변수별로 결과 저장 (테이블 + 검정값)
    row_var_results = []
    
    for row_var in row_variables:
        # 각 행 변수에 대한 교차표 생성
        if options['show_total']:
            if weights is not None:
                crosstab = pd.crosstab(df[row_var], df[col_var], values=weights, aggfunc='sum', margins=True, margins_name='Total', dropna=True).fillna(0)
            else:
                crosstab = pd.crosstab(df[row_var], df[col_var], margins=True, margins_name='Total')
        else:
            if weights is not None:
                crosstab = pd.crosstab(df[row_var], df[col_var], values=weights, aggfunc='sum', dropna=True).fillna(0)
            else:
                crosstab = pd.crosstab(df[row_var], df[col_var])
        
        # 카이제곱 검정 (백분율이 하나라도 있으면 수행)
        chi_square_result = None
        if options.get('auto_test') and (stats_config.get('row_pct') or stats_config.get('col_pct')):
            from scipy.stats import chi2_contingency
            # Total 행/열 제외하고 검정
            if options['show_total']:
                test_data = crosstab.iloc[:-1, :-1]
            else:
                test_data = crosstab
            
            try:
                chi2, p_value, dof, expected = chi2_contingency(test_data)
                row_label = get_variable_label(row_var, codebook_data) if codebook_data else row_var
                
                # 카이제곱 검정 포맷팅
                chi_format = options.get('chi_format', {
                    'show_value': True,
                    'show_df': True,
                    'show_stars': True,
                    'show_pvalue': False
                })
                formatted_stat = format_test_statistic(chi2, int(dof), p_value, chi_format, 'chi')
                
                chi_square_result = {
                    'row_var': row_var,
                    'row_label': row_label,
                    'chi2': chi2,
                    'p_value': p_value,
                    'dof': dof,
                    'test_name': 'Chi-square',
                    'formatted': formatted_stat
                }
            except Exception as e:
                import traceback
                print(f"카이제곱 검정 오류 ({row_var}): {e}")
                traceback.print_exc()
                pass
        
        # 값 레이블 적용
        if codebook_data:
            crosstab = apply_labels_to_dataframe(crosstab, row_var, col_var, codebook_data)
        
        # 행 인덱스에 변수명 추가 (계층 구조)
        row_label = get_variable_label(row_var, codebook_data) if codebook_data else row_var
        new_index = [f"{row_label}: {idx}" if idx != 'Total' else idx for idx in crosstab.index]
        crosstab.index = new_index
        
        # 이 행 변수의 테이블들
        tables = []
        
        # 빈도 테이블
        if stats_config.get('count'):
            tables.append({
                'type': 'count',
                'label': '빈도 (N)',
                'data': dataframe_to_template_dict(crosstab.copy())
            })
        
        # 행 백분율
        if stats_config.get('row_pct'):
            if options['show_total']:
                # 가중치 이미 적용됨 (crosstab 생성 시)
                row_pct = crosstab.div(crosstab.sum(axis=1), axis=0) * 100
                # Total 행이 이미 포함됨
            else:
                row_pct = crosstab.div(crosstab.sum(axis=1), axis=0) * 100
            
            # 사례수 계산 (레이블 적용 전, 원본 인덱스로)
            if weights is not None:
                # 가중치 합계가 사례수가 됨
                count_data = crosstab.copy()
            else:
                count_data = pd.crosstab(df[row_var], df[col_var])
                if options['show_total']:
                    count_data.loc['Total'] = count_data.sum()
            
            # 각 행의 사례수를 미리 계산 (인덱스 순서대로)
            row_counts_dict = {}
            for idx in count_data.index:
                val = count_data.loc[idx]
                if isinstance(val, pd.Series):
                    row_counts_dict[idx] = int(round(val.sum())) if weights is not None else int(val.sum())
                else:
                    row_counts_dict[idx] = int(round(val)) if weights is not None else int(val)
            
            print(f"\n=== 사례수 계산 디버깅 ({row_var}) ===")
            print(f"원본 인덱스: {list(count_data.index)}")
            print(f"사례수 딕셔너리: {row_counts_dict}")
            
            # 레이블 적용 전 빈도를 백분율과 결합하여 문자열로 변환
            # count_data는 crosstab과 동일한 구조
            
            # DataFrame을 순회하며 빈도(백분율) 형식으로 변환
            combined_data = []
            for i in range(len(row_pct)):
                row_vals = []
                for j in range(len(row_pct.columns)):
                    if j < len(crosstab.columns):
                        cnt = crosstab.iloc[i, j]
                        pct = row_pct.iloc[i, j]
                        row_vals.append(f"{int(round(cnt))}({pct:.1f})")
                    else:
                        # 계 열 (아직 합산 전일 수 있음)
                        pct = row_pct.iloc[i, j]
                        cnt = crosstab.iloc[i].sum()
                        row_vals.append(f"{int(round(cnt))}({pct:.1f})")
                combined_data.append(row_vals)
            
            # row_pct를 문자열 기반 DataFrame으로 재생성
            row_pct_combined = pd.DataFrame(combined_data, index=row_pct.index, columns=row_pct.columns)
            
            # 레이블 적용
            if codebook_data:
                row_pct_combined = apply_labels_to_dataframe(row_pct_combined, row_var, col_var, codebook_data)
            
            # 계 추가 (문자열이라 sum이 안 되므로 직접 계산)
            row_counts_list = []
            row_pcts_list = []
            for i in range(len(row_pct)):
                row_counts_list.append(int(round(crosstab.iloc[i].sum())))
                row_pcts_list.append(row_pct.iloc[i].sum())
            
            row_pct_combined['계'] = [f"{cnt}({pct:.1f})" for cnt, pct in zip(row_counts_list, row_pcts_list)]
            
            # 사례수 추가
            case_counts = []
            original_indices = list(crosstab.index)
            for i, labeled_idx in enumerate(row_pct_combined.index):
                if i < len(original_indices):
                    orig_idx = original_indices[i]
                    count = int(round(crosstab.loc[orig_idx].sum()))
                    case_counts.append(f"({count})")
                else:
                    case_counts.append("(0)")
            
            row_pct_combined['사례수'] = case_counts
            
            new_index = [f"{row_label}: {idx}" if idx != 'Total' else idx for idx in row_pct_combined.index]
            row_pct_combined.index = new_index
            
            tables.append({
                'type': 'row_pct',
                'label': '행 % (Row %)',
                'data': dataframe_to_template_dict(row_pct_combined)
            })
        
        # 열 백분율
        if stats_config.get('col_pct'):
            if options['show_total']:
                # 가중치 적용 col %
                if weights is not None:
                     col_pct = crosstab.div(crosstab.sum(axis=0), axis=1) * 100
                else:
                    col_pct = pd.crosstab(df[row_var], df[col_var], normalize='columns') * 100
                    total_col = pd.crosstab(df[row_var], df[col_var], normalize='all').sum(axis=1) * 100
                    col_pct['Total'] = total_col
            else:
                if weights is not None:
                    col_pct = crosstab.div(crosstab.sum(axis=0), axis=1) * 100
                else:
                    col_pct = pd.crosstab(df[row_var], df[col_var], normalize='columns') * 100
            
            if codebook_data:
                col_pct = apply_labels_to_dataframe(col_pct, row_var, col_var, codebook_data)
            
            new_index = [f"{row_label}: {idx}" if idx != 'Total' else idx for idx in col_pct.index]
            col_pct.index = new_index
            
            tables.append({
                'type': 'col_pct',
                'label': '열 % (Column %)',
                'data': dataframe_to_template_dict(col_pct)
            })
        
        # 이 행 변수의 결과 저장
        row_var_results.append({
            'row_var': row_var,
            'row_label': row_label,
            'tables': tables,
            'chi_square': chi_square_result
        })
    
    return {'row_var_results': row_var_results}

def perform_frequency_analysis(df, row_var, col_var, stats_config, codebook_data, options):
    """빈도 분석 수행"""
    # 교차표 생성
    if options['show_total']:
        crosstab = pd.crosstab(df[row_var], df[col_var], margins=True, margins_name='Total')
    else:
        crosstab = pd.crosstab(df[row_var], df[col_var])
    
    # 값 레이블 적용 (DataFrame 상태에서)
    if codebook_data:
        crosstab = apply_labels_to_dataframe(crosstab, row_var, col_var, codebook_data)
    
    result = {
        'tables': []
    }
    
    # 빈도 테이블
    if stats_config.get('count'):
        result['tables'].append({
            'type': 'count',
            'label': '빈도 (N)',
            'data': dataframe_to_template_dict(crosstab)
        })
    
    # 행 백분율
    if stats_config.get('row_pct'):
        if options['show_total']:
            row_pct = pd.crosstab(df[row_var], df[col_var], normalize='index') * 100
            # Total 행 추가
            total_row = pd.crosstab(df[row_var], df[col_var], normalize='all').sum(axis=0) * 100
            row_pct.loc['Total'] = total_row
        else:
            row_pct = pd.crosstab(df[row_var], df[col_var], normalize='index') * 100
        
        # 계 및 사례수 컬럼 추가를 위해 데이터 보강
        # DataFrame을 순회하며 빈도(백분율) 형식으로 변환
        combined_data = []
        original_crosstab = crosstab.copy()
        
        for i in range(len(row_pct)):
            row_vals = []
            for j in range(len(row_pct.columns)):
                cnt = original_crosstab.iloc[i, j]
                pct = row_pct.iloc[i, j]
                row_vals.append(f"{int(round(cnt))}({pct:.1f})")
            combined_data.append(row_vals)
            
        row_pct_combined = pd.DataFrame(combined_data, index=row_pct.index, columns=row_pct.columns)
        
        # 계 컬럼 추가
        row_counts = original_crosstab.sum(axis=1)
        row_pcts = row_pct.sum(axis=1) # 보통 100.0
        row_pct_combined['계'] = [f"{int(round(cnt))}({pct:.1f})" for cnt, pct in zip(row_counts, row_pcts)]
        
        # 사례수 컬럼 추가
        row_pct_combined['사례수'] = row_counts.apply(lambda x: f'({int(round(x))})')
        
        # 값 레이블 적용
        if codebook_data:
            row_pct_combined = apply_labels_to_dataframe(row_pct_combined, row_var, col_var, codebook_data)
        
        result['tables'].append({
            'type': 'row_pct',
            'label': '행 % (Row %)',
            'data': dataframe_to_template_dict(row_pct_combined)
        })
    
    # 열 백분율
    if stats_config.get('col_pct'):
        if options['show_total']:
            col_pct = pd.crosstab(df[row_var], df[col_var], normalize='columns') * 100
            # Total 열 추가
            total_col = pd.crosstab(df[row_var], df[col_var], normalize='all').sum(axis=1) * 100
            col_pct['Total'] = total_col
        else:
            col_pct = pd.crosstab(df[row_var], df[col_var], normalize='columns') * 100
        
        # 값 레이블 적용
        if codebook_data:
            col_pct = apply_labels_to_dataframe(col_pct, row_var, col_var, codebook_data)
        
        result['tables'].append({
            'type': 'col_pct',
            'label': '열 % (Column %)',
            'data': dataframe_to_template_dict(col_pct)
        })
    
    # 카이제곱 검정 (백분율이 하나라도 있으면 수행)
    if options.get('auto_test') and (stats_config.get('row_pct') or stats_config.get('col_pct')):
        from scipy.stats import chi2_contingency
        # Total 행/열 제외하고 검정
        if options['show_total']:
            test_data = crosstab.iloc[:-1, :-1]
        else:
            test_data = crosstab
        
        try:
            # 카이제곱 검정 시 가중치 적용 여부에 따라 데이터 준비
             # perform_frequency_analysis에서는 test_data가 이미 crosstab임
            if weights is not None:
                # 가중치가 있는 경우 이미 데이터가 가중치 적용되어 있을 수 있으나,
                # perform_frequency_analysis 구조상 여기서 다시 확인하거나
                # 앞서 계산된 crosstab이 가중치 적용된 것인지 확인 필요.
                # 하지만 이 함수 내에서는 crosstab 생성 시 weights를 직접 쓰지 않았음 (수정 필요할 수도 있음)
                # NOTE: perform_frequency_analysis는 현재 구조상 앞부분에서 weights를 crosstab 생성에 쓰지 않고 있음.
                # 따라서 여기서 가중치 crosstab을 다시 만들어야 함.
                
                # Total 제외 로직과 결합해야 하므로 복잡함.
                # 우선 기존 로직은 test_data를 그대로 썼음.
                pass 

            # 가중치 반올림 (SPSS 매칭)
            test_data_rounded = test_data.round()
            chi2, p_value, dof, expected = chi2_contingency(test_data_rounded, correction=False)
            
            # 카이제곱 검정 포맷팅
            chi_format = options.get('chi_format', {
                'show_value': True,
                'show_df': True,
                'show_stars': True,
                'show_pvalue': False
            })
            formatted_stat = format_test_statistic(chi2, int(dof), p_value, chi_format, 'chi')
            
            result['chi_square'] = {
                'chi2': chi2,
                'p_value': p_value,
                'dof': dof,
                'test_name': 'Chi-square',
                'formatted': formatted_stat
            }
        except:
            result['chi_square'] = None
    
    result['percent_symbol'] = options.get('percent_symbol', True)
    return result

def perform_descriptive_analysis_combined(df, row_variables, col_var, stats_config, codebook_data, options, weight_variable=None):
    """여러 행 변수를 하나의 테이블로 통합한 기술통계 분석"""
    
    # 가중치 변수 확인
    weights = None
    if weight_variable and weight_variable in df.columns:
        weights = pd.to_numeric(df[weight_variable], errors='coerce').fillna(0)
    
    # 열 변수를 연속형으로 변환 시도
    try:
        df_copy = df.copy()
        df_copy[col_var] = pd.to_numeric(df_copy[col_var], errors='coerce')
    except:
        return {'error': f'{col_var}는 연속형 변수가 아닙니다.'}
    
    # 열 변수 레이블
    col_label = get_variable_label(col_var, codebook_data) if codebook_data else col_var
    
    # 각 통계량 계산
    stats_to_calc = []
    
    if stats_config.get('mean'):
        if weights is not None:
             stats_to_calc.append(('평균', 'mean', lambda x, w: weighted_mean(x, w)))
        else:
            stats_to_calc.append(('평균', 'mean', lambda x: x.mean()))
    if stats_config.get('median'):
        if weights is not None:
            stats_to_calc.append(('중앙값', 'median', lambda x, w: weighted_median(x, w) if hasattr(weighted_mean, '__code__') else weighted_quantile(x, w, 0.5)))
        else:
            stats_to_calc.append(('중앙값', 'median', lambda x: x.median()))
    if stats_config.get('std'):
        if weights is not None:
            stats_to_calc.append(('표준편차', 'std', lambda x, w: weighted_std(x, w)))
        else:
            stats_to_calc.append(('표준편차', 'std', lambda x: x.std()))
    if stats_config.get('min'):
        stats_to_calc.append(('최솟값', 'min', lambda x: x.min()))
    if stats_config.get('max'):
        stats_to_calc.append(('최댓값', 'max', lambda x: x.max()))
    if stats_config.get('q1'):
        if weights is not None:
            stats_to_calc.append(('Q1 (25%)', 'q1', lambda x, w: weighted_quantile(x, w, 0.25)))
        else:
            stats_to_calc.append(('Q1 (25%)', 'q1', lambda x: x.quantile(0.25)))
    if stats_config.get('q3'):
        if weights is not None:
            stats_to_calc.append(('Q3 (75%)', 'q3', lambda x, w: weighted_quantile(x, w, 0.75)))
        else:
            stats_to_calc.append(('Q3 (75%)', 'q3', lambda x: x.quantile(0.75)))
    
    # 열 헤더 생성 (통계량 레이블)
    columns = [f"{col_label} ({label})" for label, _, _ in stats_to_calc]
    
    # 각 행 변수별로 처리
    row_var_results = []
    
    for row_var in row_variables:
        grouped = df_copy.groupby(row_var)[col_var]
        row_label = get_variable_label(row_var, codebook_data) if codebook_data else row_var
        
        rows = []
        
        # 이 행 변수의 각 그룹에 대해
        for group_val in sorted(grouped.groups.keys()):
            # 값 레이블 적용
            if codebook_data and row_var in codebook_data and 'values' in codebook_data[row_var]:
                value_labels = codebook_data[row_var]['values']
                try:
                    key = str(int(float(group_val)))
                except:
                    key = str(group_val)
                label = value_labels.get(key, str(group_val))
            else:
                label = str(group_val)
            
            # 행 레이블
            row_name = f"{row_label}: {label}"
            
            # 각 통계량 계산
            values = []
            for stat_label, stat_type, func in stats_to_calc:
                subset = df_copy[df_copy[row_var] == group_val]
                target_vals = subset[col_var].dropna()
                
                if weights is not None:
                    if stat_type in ['mean', 'median', 'std', 'q1', 'q3']:
                        target_weights = weights[target_vals.index]
                        value = func(target_vals, target_weights)
                    else:
                        value = func(target_vals)
                else:
                    value = func(target_vals)
                    
                values.append(value)
            
            rows.append({
                'label': row_name,
                'values': values
            })
        
        # Total 추가
        if options['show_total']:
            total_values = []
            for stat_label, stat_type, func in stats_to_calc:
                target_vals = df_copy[col_var].dropna()
                if weights is not None:
                     if stat_type in ['mean', 'median', 'std', 'q1', 'q3']:
                        target_weights = weights[target_vals.index]
                        value = func(target_vals, target_weights)
                     else:
                        value = func(target_vals)
                else:
                    value = func(target_vals)
                total_values.append(value)
            
            rows.append({
                'label': 'Total',
                'values': total_values
            })
        
        # 통계 검정 (평균이 선택된 경우만)
        statistical_test = None
        if options.get('auto_test') and stats_config.get('mean'):
            groups = [group[col_var].dropna() for name, group in df_copy.groupby(row_var)]
            
            # t/F 검정 포맷팅
            tf_format = options.get('tf_format', {
                'show_label': True,
                'show_value': True,
                'show_df': True,
                'show_stars': True,
                'show_pvalue': False
            })
            
            # 그룹이 2개인 경우 t-test
            if len(groups) == 2:
                try:
                    from scipy import stats as sp_stats
                    
                    if weights is not None:
                         # 가중치 t-test 수행
                         
                         # 리스트 준비
                         w1 = weights[groups[0].index]
                         w2 = weights[groups[1].index]
                         
                         # 1. 가중 Levene 검정
                         levene_formatted_str = ""
                         
                         levene_result = weighted_levene(groups, [w1, w2])
                         if levene_result:
                             f_levene, p_levene, df1_levene, df2_levene = levene_result
                             levene_stat, levene_p = f_levene, p_levene
                             # Levene 결과 포맷팅
                             levene_formatted_str = format_test_statistic(f_levene, (df1_levene, df2_levene), p_levene, tf_format, 'f')
                         else:
                             levene_stat, levene_p = 0, 1 # 실패 시 등분산 가정
                             
                         # 2. 가중 t-test (Welch)
                         t_welch, p_welch, df_welch_val = weighted_ttest(groups[0], groups[1], w1, w2)
                         
                         # 3. 가중 t-test (Pooled/Equal Variance)
                         t_pooled, p_pooled, df_pooled_val = weighted_ttest_pooled(groups[0], groups[1], w1, w2)
                         
                         # 결과 매핑
                         t_stat_equal, p_value_equal, df_equal = t_pooled, p_pooled, df_pooled_val
                         t_stat_welch, p_value_welch, df_welch = t_welch, p_welch, df_welch_val
                         
                         # 선택 로직
                         if levene_p < 0.05:
                             t_stat, p_value, df_val = t_stat_welch, p_value_welch, df_welch
                             test_name = "t-test (Welch)"
                         else:
                             t_stat, p_value, df_val = t_stat_equal, p_value_equal, df_equal
                             test_name = "t-test"
                         
                    else:
                        # unweighted logic (unchanged)
                        from scipy import stats as sp_stats
                        
                        # Levene의 등분산 검정
                        levene_stat, levene_p = sp_stats.levene(groups[0], groups[1])
                        # Levene 포맷팅
                        levene_formatted_str = f"F={levene_stat:.3f} ({levene_p:.3f})"
                        
                        # 등분산 가정 t-test (Student's t-test)
                        t_stat_equal, p_value_equal = sp_stats.ttest_ind(groups[0], groups[1], equal_var=True)
                        df_equal = len(groups[0]) + len(groups[1]) - 2
                        
                        # 등분산 가정하지 않음 t-test (Welch's t-test)
                        t_stat_welch, p_value_welch = sp_stats.ttest_ind(groups[0], groups[1], equal_var=False)
                        # Welch's t-test의 자유도 계산
                        n1, n2 = len(groups[0]), len(groups[1])
                        v1, v2 = groups[0].var(ddof=1), groups[1].var(ddof=1)
                        df_welch = ((v1/n1 + v2/n2)**2) / ((v1/n1)**2/(n1-1) + (v2/n2)**2/(n2-1))
                        
                        # Levene 검정 결과에 따라 적절한 t-test 선택
                        if levene_p < 0.05:
                            t_stat, p_value, df_val = t_stat_welch, p_value_welch, df_welch
                            test_name = "t-test (Welch)"
                        else:
                            t_stat, p_value, df_val = t_stat_equal, p_value_equal, df_equal
                            test_name = "t-test"
                    
                    formatted_stat = format_test_statistic(t_stat, df_val, p_value, tf_format, 't')
                    
                    statistical_test = {
                        'row_var': row_var,
                        'row_label': row_label,
                        't_statistic': t_stat,
                        'p_value': p_value,
                        'test_name': test_name,
                        'df': df_val,
                        'levene_statistic': levene_stat,
                        'levene_p': levene_p,
                        't_equal': t_stat_equal,
                        'p_equal': p_value_equal,
                        'df_equal': df_equal,
                        't_welch': t_stat_welch,
                        'p_welch': p_value_welch,
                        'df_welch': df_welch,
                        'formatted': formatted_stat,
                        'levene_formatted': levene_formatted_str
                    }
                except:
                    pass
            
            # 그룹이 3개 이상인 경우 ANOVA
            elif len(groups) >= 3:
                try:
                    if weights is not None:
                         # 가중치 ANOVA
                         weights_list = [weights[g.index] for g in groups]
                         anova_result = weighted_anova(groups, weights_list)
                         
                         if anova_result:
                             f_stat, p_value, df_between, df_within = anova_result
                         else:
                             raise ValueError("Weighted ANOVA failed")
                    else:
                        from scipy import stats as sp_stats
                        f_stat, p_value = sp_stats.f_oneway(*groups)
                        df_between = len(groups) - 1
                        df_within = sum(len(g) for g in groups) - len(groups)
                    
                    formatted_stat = format_test_statistic(f_stat, (df_between, df_within), p_value, tf_format, 'f')
                    
                    statistical_test = {
                        'row_var': row_var,
                        'row_label': row_label,
                        'f_statistic': f_stat,
                        'p_value': p_value,
                        'test_name': 'ANOVA (F-test)',
                        'df_between': df_between,
                        'df_within': df_within,
                        'formatted': formatted_stat
                    }
                except:
                    pass
        
        # 이 행 변수의 결과 저장
        row_var_results.append({
            'rows': rows,
            'statistical_test': statistical_test
        })
    
    return {
        'columns': columns,
        'row_var_results': row_var_results
    }

def perform_descriptive_analysis(df, row_var, col_var, stats_config, codebook_data, options):
    """기술통계 분석 수행"""
    # 열 변수를 연속형으로 변환 시도
    try:
        df_copy = df.copy()
        df_copy[col_var] = pd.to_numeric(df_copy[col_var], errors='coerce')
    except:
        return {'error': f'{col_var}는 연속형 변수가 아닙니다.'}
    
    # 그룹별 기술통계
    grouped = df_copy.groupby(row_var)[col_var]
    
    result = {
        'statistics': []
    }
    
    # 각 통계량 계산
    stats_to_calc = []
    
    if stats_config.get('mean'):
        stats_to_calc.append(('평균', 'mean', lambda x: x.mean()))
    if stats_config.get('median'):
        stats_to_calc.append(('중앙값', 'median', lambda x: x.median()))
    if stats_config.get('std'):
        stats_to_calc.append(('표준편차', 'std', lambda x: x.std()))
    if stats_config.get('min'):
        stats_to_calc.append(('최솟값', 'min', lambda x: x.min()))
    if stats_config.get('max'):
        stats_to_calc.append(('최댓값', 'max', lambda x: x.max()))
    if stats_config.get('q1'):
        stats_to_calc.append(('Q1 (25%)', 'q1', lambda x: x.quantile(0.25)))
    if stats_config.get('q3'):
        stats_to_calc.append(('Q3 (75%)', 'q3', lambda x: x.quantile(0.75)))
    
    # 통계량 계산 및 테이블 생성
    for label, stat_type, func in stats_to_calc:
        stat_values = grouped.apply(func)
        
        # Total 계산
        if options['show_total']:
            total_value = func(df_copy[col_var].dropna())
            stat_values['Total'] = total_value
        
        # 값 레이블 적용
        if codebook_data and row_var in codebook_data and 'values' in codebook_data[row_var]:
            value_labels = codebook_data[row_var]['values']
            new_index = []
            for idx in stat_values.index:
                if idx == 'Total':
                    new_index.append('Total')
                else:
                    # 정수로 변환 시도
                    try:
                        key = str(int(float(idx)))
                    except:
                        key = str(idx)
                    new_index.append(value_labels.get(key, str(idx)))
            stat_values.index = new_index
        
        # 값 소수점 포맷 적용
        mean_decimals = options.get('mean_decimals', 2)
        formatted_values = []
        for val in stat_values.values:
            if stat_type in ['mean', 'median', 'std', 'q1', 'q3']:
                formatted_values.append(round(val, mean_decimals))
            else:
                formatted_values.append(val)
        
        result['statistics'].append({
            'type': stat_type,
            'label': label,
            'data': {
                'index': list(stat_values.index),
                'values': formatted_values
            }
        })
    
    # 통계 검정 (평균이 선택된 경우)
    if options.get('auto_test') and stats_config.get('mean'):
        groups = [group[col_var].dropna() for name, group in df_copy.groupby(row_var)]
        
        # t/F 검정 포맷팅
        tf_format = options.get('tf_format', {
            'show_value': True,
            'show_df': True,
            'show_stars': True,
            'show_pvalue': False
        })
        
        # 그룹이 2개인 경우 t-test
        if len(groups) == 2:
            try:
                t_stat, p_value = stats.ttest_ind(groups[0], groups[1])
                df_val = len(groups[0]) + len(groups[1]) - 2
                
                formatted_stat = format_test_statistic(t_stat, df_val, p_value, tf_format, 't')
                
                result['statistical_test'] = {
                    't_statistic': t_stat,
                    'p_value': p_value,
                    'test_name': 't-test',
                    'df': df_val,
                    'formatted': formatted_stat
                }
            except:
                result['statistical_test'] = None
        
        # 그룹이 3개 이상인 경우 ANOVA
        elif len(groups) >= 3:
            try:
                f_stat, p_value = stats.f_oneway(*groups)
                df_between = len(groups) - 1
                df_within = sum(len(g) for g in groups) - len(groups)
                
                formatted_stat = format_test_statistic(f_stat, (df_between, df_within), p_value, tf_format, 'f')
                
                result['statistical_test'] = {
                    'f_statistic': f_stat,
                    'p_value': p_value,
                    'test_name': 'ANOVA (F-test)',
                    'df_between': df_between,
                    'df_within': df_within,
                    'formatted': formatted_stat
                }
            except:
                result['statistical_test'] = None
    
    return result

def apply_labels_to_dataframe(df, row_var, col_var, codebook_data):
    """DataFrame의 인덱스와 컬럼에 값 레이블 적용"""
    df_copy = df.copy()
    
    print(f"=== 레이블 적용 시작 ===")
    print(f"행 변수: {row_var}, 열 변수: {col_var}")
    print(f"코드북에 {row_var} 있음: {row_var in codebook_data}")
    print(f"코드북에 {col_var} 있음: {col_var in codebook_data}")
    
    # 행 인덱스 레이블 변경
    if row_var in codebook_data and 'values' in codebook_data[row_var]:
        value_labels = codebook_data[row_var]['values']
        print(f"{row_var}의 값 레이블: {value_labels}")
        print(f"원본 인덱스: {list(df_copy.index)}")
        
        new_index = []
        for idx in df_copy.index:
            if idx == 'Total':
                new_index.append('Total')
            else:
                # 정수로 변환 시도
                try:
                    key = str(int(float(idx)))
                except:
                    key = str(idx)
                label = value_labels.get(key, str(idx))
                print(f"  인덱스 {idx} -> 키 {key} -> 레이블 {label}")
                new_index.append(label)
        df_copy.index = new_index
        print(f"변경된 인덱스: {list(df_copy.index)}")
    
    # 열 인덱스 레이블 변경
    if col_var in codebook_data and 'values' in codebook_data[col_var]:
        value_labels = codebook_data[col_var]['values']
        print(f"{col_var}의 값 레이블: {value_labels}")
        print(f"원본 컬럼: {list(df_copy.columns)}")
        
        new_columns = []
        for col in df_copy.columns:
            if col == 'Total':
                new_columns.append('Total')
            else:
                # 정수로 변환 시도
                try:
                    key = str(int(float(col)))
                except:
                    key = str(col)
                label = value_labels.get(key, str(col))
                print(f"  컬럼 {col} -> 키 {key} -> 레이블 {label}")
                new_columns.append(label)
        df_copy.columns = new_columns
        print(f"변경된 컬럼: {list(df_copy.columns)}")
    
    return df_copy

def apply_value_labels_to_result(result, row_var, col_var, codebook_data):
    """빈도 분석 결과에 값 레이블 적용"""
    if not codebook_data:
        return result
    
    for table in result.get('tables', []):
        data_dict = table['data']
        
        # 행 인덱스 레이블 변경
        if row_var in codebook_data and 'values' in codebook_data[row_var]:
            value_labels = codebook_data[row_var]['values']
            new_index = [value_labels.get(str(idx), idx) for idx in data_dict['index']]
            data_dict['index'] = new_index
        
        # 열 인덱스 레이블 변경
        if col_var in codebook_data and 'values' in codebook_data[col_var]:
            value_labels = codebook_data[col_var]['values']
            new_columns = [value_labels.get(str(col), col) for col in data_dict['columns']]
            data_dict['columns'] = new_columns
    
    return result

def apply_value_labels_to_desc_result(result, row_var, codebook_data):
    """기술통계 결과에 값 레이블 적용"""
    if not codebook_data:
        return result
    
    if row_var in codebook_data and 'values' in codebook_data[row_var]:
        value_labels = codebook_data[row_var]['values']
        
        for stat in result.get('statistics', []):
            data_dict = stat['data']
            new_index = [value_labels.get(str(idx), idx) for idx in data_dict['index']]
            data_dict['index'] = new_index
    
    return result

def load_codebook(codebook):
    """코드북 파일에서 변수 및 값 레이블 로드"""
    df = pd.read_excel(codebook.file.path)
    
    codebook_dict = {}
    current_var = None
    
    for _, row in df.iterrows():
        var_name = row.get('Variable')
        var_label = row.get('Variable Label')
        value = row.get('Value')
        value_label = row.get('Value Label')
        
        # 변수명이 있으면 새 변수 시작
        if pd.notna(var_name) and str(var_name).strip():
            current_var = str(var_name).strip()
            if current_var not in codebook_dict:
                codebook_dict[current_var] = {
                    'label': str(var_label).strip() if pd.notna(var_label) else current_var,
                    'values': {}
                }
        
        # 값과 값 레이블 추가 (현재 변수가 있고, 값이 있을 때)
        # Variable과 Variable Label이 같은 행에 Value도 있을 수 있음
        if current_var and pd.notna(value) and pd.notna(value_label):
            # 값을 정수로 변환 시도, 실패하면 문자열로
            try:
                value_key = str(int(float(value)))
            except:
                value_key = str(value).strip()
            
            codebook_dict[current_var]['values'][value_key] = str(value_label).strip()
    
    # 디버그 출력
    print("=== 코드북 로드 완료 ===")
    for var, info in list(codebook_dict.items())[:5]:  # 처음 5개만 출력
        print(f"변수: {var}")
        print(f"  레이블: {info['label']}")
        print(f"  값들: {info['values']}")
    
    return codebook_dict

def get_variable_label(var_name, codebook_data):
    """변수명에 대한 레이블 반환"""
    if codebook_data and var_name in codebook_data:
        return codebook_data[var_name]['label']
    return var_name

@login_required
def get_codebooks_for_dataset(request, dataset_id):
    """특정 데이터셋에 연결된 코드북 목록 반환"""
    dataset = get_object_or_404(SurveyData, id=dataset_id, user=request.user)
    
    codebooks = Codebook.objects.filter(
        user=request.user,
        dataset=dataset
    ).values('id', 'name')
    
    return JsonResponse({
        'codebooks': list(codebooks)
    })

@login_required
def save_analysis_preset(request):
    """분석 프리셋 저장"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=400)
    
    try:
        dataset_id = request.POST.get('dataset_id')
        name = request.POST.get('name')
        config_json = request.POST.get('configuration')
        
        if not dataset_id or not name or not config_json:
            return JsonResponse({'error': 'Missing required fields'}, status=400)
            
        dataset = get_object_or_404(SurveyData, id=dataset_id, user=request.user)
        configuration = json.loads(config_json)
        
        preset = AnalysisPreset.objects.create(
            user=request.user,
            dataset=dataset,
            name=name,
            configuration=configuration
        )
        
        return JsonResponse({
            'success': True,
            'preset_id': preset.id,
            'message': '프리셋이 저장되었습니다.'
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def get_analysis_presets(request, dataset_id):
    """특정 데이터셋의 프리셋 목록 조회"""
    dataset = get_object_or_404(SurveyData, id=dataset_id, user=request.user)
    
    presets = AnalysisPreset.objects.filter(
        user=request.user, 
        dataset=dataset
    ).values('id', 'name', 'created_at')
    
    return JsonResponse({
        'presets': list(presets)
    }, encoder=DjangoJSONEncoder)

@login_required
def load_analysis_preset(request, preset_id):
    """프리셋 설정 로드"""
    preset = get_object_or_404(AnalysisPreset, id=preset_id, user=request.user)
    
    return JsonResponse({
        'configuration': preset.configuration
    })

@login_required
def delete_analysis_preset(request, preset_id):
    """프리셋 삭제"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=400)
        
    preset = get_object_or_404(AnalysisPreset, id=preset_id, user=request.user)
    preset.delete()
    
    return JsonResponse({
        'success': True,
        'message': '프리셋이 삭제되었습니다.'
    })

@login_required
def get_variables_for_dataset(request, dataset_id):
    """특정 데이터셋의 변수 목록 반환"""
    dataset = get_object_or_404(SurveyData, id=dataset_id, user=request.user)
    
    try:
        if dataset.file.name.lower().endswith('.csv'):
            df = pd.read_csv(dataset.file.path)
        elif dataset.file.name.lower().endswith('.sav'):
            df = pd.read_spss(dataset.file.path, convert_categoricals=False)
        else:
            df = pd.read_excel(dataset.file.path)
        
        variables = df.columns.tolist()
        
        return JsonResponse({
            'variables': variables,
            'count': len(variables)
        })
    except Exception as e:
        return JsonResponse({
            'error': str(e)
        }, status=500)

@login_required
def get_variable_values(request, dataset_id, variable):
    """특정 변수의 고유값 및 빈도 반환"""
    dataset = get_object_or_404(SurveyData, id=dataset_id, user=request.user)
    
    try:
        if dataset.file.name.endswith('.csv'):
            df = pd.read_csv(dataset.file.path)
        else:
            df = pd.read_excel(dataset.file.path)
        
        if variable not in df.columns:
            return JsonResponse({'error': 'Variable not found'}, status=404)
        
        value_counts = df[variable].value_counts().sort_index()
        
        values = []
        for val, count in value_counts.items():
            values.append({
                'value': str(val),
                'count': int(count),
                'percentage': round(count / len(df) * 100, 1)
            })
        
        return JsonResponse({
            'values': values,
            'total': len(df)
        })
    except Exception as e:
        return JsonResponse({
            'error': str(e)
        }, status=500)

@login_required
def recode_variable(request):
    """변수 변환 메인 페이지"""
    from surveys.models import RecodeRule
    
    datasets = SurveyData.objects.filter(user=request.user)
    existing_rules = RecodeRule.objects.filter(user=request.user).select_related('dataset')
    
    return render(request, 'analysis/recode.html', {
        'datasets': datasets,
        'existing_rules': existing_rules
    })

@login_required
def recode_preview(request):
    """변환 미리보기 (AJAX)"""
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=400)
    
    try:
        dataset_id = request.POST.get('dataset_id')
        source_variable = request.POST.get('source_variable')
        rules_json = request.POST.get('rules')
        
        dataset = get_object_or_404(SurveyData, id=dataset_id, user=request.user)
        
        # 데이터 로드
        if dataset.file.name.endswith('.csv'):
            df = pd.read_csv(dataset.file.path)
        else:
            df = pd.read_excel(dataset.file.path)
        
        if source_variable not in df.columns:
            return JsonResponse({'error': 'Variable not found'}, status=404)
        
        # 규칙 파싱
        import json
        rules = json.loads(rules_json)
        
        # 변환 적용
        original_values = df[source_variable].copy()
        new_values = apply_recode_rules(original_values, rules)
        
        # 미리보기 데이터 생성
        preview_data = []
        
        # 원본 값별로 그룹화
        for orig_val in original_values.dropna().unique():
            mask = original_values == orig_val
            new_val = new_values[mask].iloc[0] if mask.any() else orig_val
            count = mask.sum()
            
            preview_data.append({
                'original': str(orig_val),
                'new': str(new_val),
                'count': int(count),
                'changed': str(orig_val) != str(new_val)
            })
        
        # 새 값별 빈도
        new_value_counts = {}
        for item in preview_data:
            new_val = item['new']
            if new_val not in new_value_counts:
                new_value_counts[new_val] = 0
            new_value_counts[new_val] += item['count']
        
        return JsonResponse({
            'success': True,
            'preview': preview_data,
            'new_value_counts': new_value_counts,
            'total': len(df)
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def recode_apply(request):
    """변환 규칙 저장 및 적용"""
    from surveys.models import RecodeRule
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=400)
    
    try:
        dataset_id = request.POST.get('dataset_id')
        source_variable = request.POST.get('source_variable')
        target_variable = request.POST.get('target_variable')
        rule_name = request.POST.get('rule_name')
        rules_json = request.POST.get('rules')
        description = request.POST.get('description', '')
        
        dataset = get_object_or_404(SurveyData, id=dataset_id, user=request.user)
        
        # 규칙 파싱
        import json
        rules = json.loads(rules_json)
        
        # 규칙 저장
        recode_rule, created = RecodeRule.objects.update_or_create(
            user=request.user,
            dataset=dataset,
            target_variable=target_variable,
            defaults={
                'name': rule_name,
                'source_variable': source_variable,
                'rules': rules,
                'description': description
            }
        )
        
        return JsonResponse({
            'success': True,
            'message': '변환 규칙이 저장되었습니다.',
            'rule_id': recode_rule.id,
            'created': created
        })
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
def recode_delete(request, rule_id):
    """변환 규칙 삭제"""
    from surveys.models import RecodeRule
    
    rule = get_object_or_404(RecodeRule, id=rule_id, user=request.user)
    rule.delete()
    
    return JsonResponse({
        'success': True,
        'message': '변환 규칙이 삭제되었습니다.'
    })

def apply_recode_rules(series, rules):
    """변환 규칙을 시리즈에 적용"""
    result = series.copy()
    
    for rule in rules:
        rule_type = rule.get('type')
        
        if rule_type == 'single':
            # 단일값 변환
            old_value = rule.get('old')
            new_value = rule.get('new')
            
            # 타입 변환 시도
            try:
                if series.dtype in ['int64', 'float64']:
                    old_value = float(old_value)
            except:
                pass
            
            mask = series == old_value
            result[mask] = new_value
        
        elif rule_type == 'range':
            # 범위 변환
            min_val = float(rule.get('min'))
            max_val = float(rule.get('max'))
            new_value = rule.get('new')
            
            mask = (series >= min_val) & (series <= max_val)
            result[mask] = new_value
        
        elif rule_type == 'condition':
            # 조건 변환
            operator = rule.get('operator')
            value = float(rule.get('value'))
            new_value = rule.get('new')
            
            if operator == '>=':
                mask = series >= value
            elif operator == '>':
                mask = series > value
            elif operator == '<=':
                mask = series <= value
            elif operator == '<':
                mask = series < value
            elif operator == '==':
                mask = series == value
            elif operator == '!=':
                mask = series != value
            else:
                continue
            
            result[mask] = new_value
        
        elif rule_type == 'else':
            # ELSE 처리 - 아직 변환되지 않은 값들
            new_value = rule.get('new')
            unchanged_mask = result == series
            result[unchanged_mask] = new_value
    
    return result

@login_required
def export_analysis_to_excel(request):
    """분석 결과를 엑셀로 내보내기"""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.http import HttpResponse
    
    # POST 데이터 파싱
    dataset_id = request.POST.get('dataset_id')
    codebook_id = request.POST.get('codebook_id')
    
    # 빈 문자열을 None으로 변환
    if not dataset_id or dataset_id == '':
        return HttpResponse("데이터셋 ID가 필요합니다.", status=400)
    
    if not codebook_id or codebook_id == '':
        codebook_id = None
    
    row_variables = request.POST.getlist('row_variables[]')
    col_variables = request.POST.getlist('col_variables[]')
    weight_variable = request.POST.get('weight_variable')
    
    # 디버깅: 받은 데이터 확인
    print(f"dataset_id: {dataset_id}")
    print(f"codebook_id: {codebook_id}")
    print(f"row_variables: {row_variables}")
    print(f"col_variables: {col_variables}")
    print(f"weight_variable: {weight_variable}")
    
    # 통계량 설정 파싱
    col_statistics = {}
    for key, value in request.POST.items():
        if key.startswith('col_statistics'):
            parts = key.split('[')
            if len(parts) >= 3:
                col_var = parts[1].rstrip(']')
                stat_type = parts[2].rstrip(']')
                if col_var not in col_statistics:
                    col_statistics[col_var] = {}
                col_statistics[col_var][stat_type] = value == 'on'
    
    # 옵션
    show_total = request.POST.get('show_total') == 'on'
    auto_test = request.POST.get('auto_test') == 'on'
    
    # 분석 실행
    user = request.user
    
    options = {
        'show_total': show_total,
        'auto_test': auto_test,
        'percent_symbol': request.POST.get('percent_symbol', 'yes') == 'yes',
        'mean_decimals': int(request.POST.get('mean_decimals', '2')),
        'chi_format': {
            'show_label': request.POST.get('show_chi_label') == 'on',
            'show_value': request.POST.get('show_chi_value') == 'on',
            'show_df': request.POST.get('show_chi_df') == 'on',
            'show_stars': request.POST.get('show_chi_stars') == 'on',
            'show_pvalue': request.POST.get('show_chi_pvalue') == 'on',
        },
        'tf_format': {
            'show_label': request.POST.get('show_tf_label') == 'on',
            'show_value': request.POST.get('show_tf_value') == 'on',
            'show_df': request.POST.get('show_tf_df') == 'on',
            'show_stars': request.POST.get('show_tf_stars') == 'on',
            'show_pvalue': request.POST.get('show_tf_pvalue') == 'on',
        }
    }
    
    results = perform_unified_analysis(
        dataset_id=dataset_id,
        codebook_id=codebook_id,
        row_variables=row_variables,
        col_variables=col_variables,
        col_statistics=col_statistics,
        user=user,
        options=options,
        weight_variable=weight_variable
    )
    
    # 엑셀 워크북 생성
    wb = Workbook()
    wb.remove(wb.active)  # 기본 시트 제거
    
    # 스타일 정의
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 각 분석 결과를 별도 시트로
    for idx, analysis in enumerate(results, 1):
        sheet_name = f"분석{idx}"
        ws = wb.create_sheet(title=sheet_name)
        
        row_num = 1
        
        # 제목
        ws.merge_cells(f'A{row_num}:E{row_num}')
        title_cell = ws[f'A{row_num}']
        title_text = f"{', '.join(analysis['row_labels'])} × {analysis['col_label']}"
        title_cell.value = title_text
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        row_num += 2
        
        # 통합 분석 (빈도 + 기술통계)
        if 'unified' in analysis and analysis['unified']:
            unified = analysis['unified']
            
            # 헤더
            ws.cell(row=row_num, column=1, value='통계량')
            for col_idx, col_name in enumerate(unified['columns'], 2):
                ws.cell(row=row_num, column=col_idx, value=col_name)
            
            # 헤더 스타일
            for col_idx in range(1, len(unified['columns']) + 2):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.fill = header_fill
                cell.font = Font(bold=True)
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            
            row_num += 1
            
            # 데이터 및 검정값
            for row_var_result in unified['row_var_results']:
                for row_data in row_var_result['rows']:
                    ws.cell(row=row_num, column=1, value=row_data['label'])
                    for col_idx, value in enumerate(row_data['values'], 2):
                        if isinstance(value, str):
                            ws.cell(row=row_num, column=col_idx, value=value)
                        else:
                            ws.cell(row=row_num, column=col_idx, value=float(value))
                    
                    # 테두리 적용
                    for col_idx in range(1, len(unified['columns']) + 2):
                        ws.cell(row=row_num, column=col_idx).border = border
                    
                    row_num += 1
                
                # 검정 통계량
                ws.cell(row=row_num, column=1, value='검정 통계량')
                
                # Chi-square (formatted 값 사용)
                chi_text = ''
                if row_var_result.get('chi_square_test') and row_var_result['chi_square_test'].get('formatted'):
                    chi_text = row_var_result['chi_square_test']['formatted']
                ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=1+unified['freq_col_count'])
                ws.cell(row=row_num, column=2, value=chi_text)
                
                # t-test/ANOVA (formatted 값 사용)
                desc_text = ''
                if row_var_result.get('descriptive_test') and row_var_result['descriptive_test'].get('formatted'):
                    desc_text = row_var_result['descriptive_test']['formatted']
                ws.merge_cells(start_row=row_num, start_column=2+unified['freq_col_count'], end_row=row_num, end_column=1+len(unified['columns']))
                ws.cell(row=row_num, column=2+unified['freq_col_count'], value=desc_text)
                
                # 테두리 적용
                for col_idx in range(1, len(unified['columns']) + 2):
                    ws.cell(row=row_num, column=col_idx).border = border
                
                row_num += 2
        
        # 빈도 분석만
        elif 'frequency' in analysis and analysis['frequency']:
            freq = analysis['frequency']
            
            for row_var_result in freq['row_var_results']:
                for table in row_var_result['tables']:
                    # 테이블 제목
                    ws.cell(row=row_num, column=1, value=table['label'])
                    ws.cell(row=row_num, column=1).font = Font(bold=True)
                    row_num += 1
                    
                    # 헤더
                    ws.cell(row=row_num, column=1, value='')
                    for col_idx, col_name in enumerate(table['data']['columns'], 2):
                        ws.cell(row=row_num, column=col_idx, value=col_name)
                        ws.cell(row=row_num, column=col_idx).fill = header_fill
                        ws.cell(row=row_num, column=col_idx).font = Font(bold=True)
                        ws.cell(row=row_num, column=col_idx).border = border
                    row_num += 1
                    
                    # 데이터
                    for row_idx, row_label in enumerate(table['data']['index']):
                        ws.cell(row=row_num, column=1, value=row_label)
                        for col_idx, value in enumerate(table['data']['data'][row_idx], 2):
                            ws.cell(row=row_num, column=col_idx, value=value)
                        
                        for col_idx in range(1, len(table['data']['columns']) + 2):
                            ws.cell(row=row_num, column=col_idx).border = border
                        row_num += 1
                    
                    row_num += 1
                
                # Chi-square (formatted 값 사용)
                if row_var_result.get('chi_square') and row_var_result['chi_square'].get('formatted'):
                    chi_text = row_var_result['chi_square']['formatted']
                    ws.cell(row=row_num, column=1, value=chi_text)
                    ws.cell(row=row_num, column=1).font = Font(italic=True)
                    row_num += 2
        
        # 기술통계만
        elif 'descriptive' in analysis and analysis['descriptive']:
            desc = analysis['descriptive']
            
            if 'error' in desc:
                ws.cell(row=row_num, column=1, value=desc['error'])
            else:
                for row_var_result in desc['row_var_results']:
                    # 헤더
                    ws.cell(row=row_num, column=1, value='통계량')
                    for col_idx, col_name in enumerate(row_var_result['statistics'][0]['data']['index'], 2):
                        ws.cell(row=row_num, column=col_idx, value=col_name)
                        ws.cell(row=row_num, column=col_idx).fill = header_fill
                        ws.cell(row=row_num, column=col_idx).font = Font(bold=True)
                        ws.cell(row=row_num, column=col_idx).border = border
                    row_num += 1
                    
                    # 데이터
                    for stat in row_var_result['statistics']:
                        ws.cell(row=row_num, column=1, value=stat['label'])
                        for col_idx, value in enumerate(stat['data']['values'], 2):
                            ws.cell(row=row_num, column=col_idx, value=value)
                        
                        for col_idx in range(1, len(stat['data']['index']) + 2):
                            ws.cell(row=row_num, column=col_idx).border = border
                        row_num += 1
                    
                    # 검정값 (formatted 값 사용)
                    if row_var_result.get('statistical_test') and row_var_result['statistical_test'].get('formatted'):
                        test_text = row_var_result['statistical_test']['formatted']
                        ws.cell(row=row_num, column=1, value=test_text)
                        ws.cell(row=row_num, column=1).font = Font(italic=True)
                        row_num += 2
        
        # 열 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    # 병합된 셀 건너뛰기
                    if hasattr(cell, 'column_letter'):
                        if column_letter is None:
                            column_letter = cell.column_letter
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    # 엑셀 파일 저장
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=analysis_results.xlsx'
    
    return response
